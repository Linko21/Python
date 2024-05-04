import json
import os

from openpyxl import Workbook
import requests
from openpyxl.utils.cell import column_index_from_string

keys = []
wb = Workbook()
ws = wb.active

url = 'https://api.lorcana-api.com/cards/all'
url2 = 'https://api.lorcana-api.com/sets/all'
param = dict()
resp = requests.get(url=url, params=param)
data = resp.json()

resp2 = requests.get(url=url2, params=param)
data2 = resp2.json()

keys = set()
card_sets = dict()


def worksheet_name(card, sets: dict):
    card_set_id = card["Set_ID"]
    nr = sets.get(card_set_id)
    return card_set_id + " " + str(nr)


for a in range(len(data2)):
    sub_obj = data2[a]
    card_sets.update({sub_obj['Set_ID']: sub_obj['Set_Num']})
key = sorted(keys)

for set_id, num in card_sets.items():
    print(set_id, num)
    # ws.title = set_id + ' ' + str(num)
    ws = wb.create_sheet(set_id + ' ' + str(num))
    for f in range(len(data)):
        sub_obj1 = data[f]
        keys.update((sub_obj1.keys()))

        key = sorted(keys)

        for k in range(len(key)):
            ws.cell(row=(0 + 1), column=(k + 1), value=key[k]);

for p in range(len(data)):
    sub_obj2 = data[p]
    sheet_name = worksheet_name(sub_obj2, card_sets)
    print("use sheet " + sheet_name + " for " + str(
        sub_obj2))  # es wird immer das sheet "INK 3" genommen weil set_id immer das letzte element von der for loop davor is
    ws = wb[sheet_name]

    for j in range(len(key)):
        key_entry = key[j]
        if key_entry in sub_obj2:
            v = sub_obj2[key_entry]
            ws.cell(row=(p + 2), column=(j + 1), value=v)

    # for j in range(len(key)):
    # key_entry = key[j]
    # if key_entry in sub_obj2:
    # v = sub_obj2[key_entry]
    # ws.cell(row = (p + 2), column = (j + 1), value = v);

wb.remove(wb['Sheet'])

wb.save('Test.xlsx')
print('saved: ' + os.path.abspath('Test.xlsx'))
