import os.path
import tkinter as tk

from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook

from orders import Utils
from orders.Utils import Dishes

from datetime import date

random_values = Utils.RandomValues(1)


def create_result():
    #print("Berechne Ergebnis tut gerade noch nix ...")
    aktuelles_Datum = date.today()
    Datum =aktuelles_Datum.strftime('%m-%d')
    wb = Workbook()
    wb = load_workbook("prices.xlsx")
    Preise = wb['Preise']
    wb = load_workbook('orders.xlsx')
    Bestellung = wb['Bestellungen']
    #print(Datum)
    #print(Geburtstdatum[4:10])
    for a in range (2,Preise.max_row+1):
        Gericht = Preise['A'+str(a)].value
        print(Gericht)
    
    for b in range (2,Bestellung.max_row+1):
        Gericht_Bestellung = Bestellung['D'+str(b)].value
        #print(Gericht_Bestellung)
    
    
        
        if Gericht == Gericht_Bestellung:
            Preis = Preise['B'+str(a)].value * Bestellung['E'+str(b)].value
            Bestellung['F1'].value = 'Preis in €'
            Bestellung['F'+str(b)].value = str(round(Preis,2))
            #print(Preis)
    #wb.save('orders.xlsx')
     
    for g in range (2,Bestellung.max_row+1):
        Geburtstdatum = Bestellung['C'+str(g)].value
        #Geburtstdatum = Geburtstdatum.strftime('%d.%m.%y')
        #print(Geburtstdatum[5:10])       
            
        if Datum == Geburtstdatum[5:10]:
            neuer_Preis = float(Bestellung['F'+str(g)].value) - float(Bestellung['F'+str(g)].value) * 0.20
            Bestellung['G1'].value = 'Preis in € mit 20% Rabatt'
            Bestellung['G'+str(g)].value = str(round(neuer_Preis,2))
            #print('Hier steht der neue Preis'+str(neuer_Preis))

    wb.save('orders.xlsx')
    print('upated: ' + os.path.abspath('orders.xlsx')) 
 
    
    # TODO lies beide excel tabellen ein und bestimme den preis pro person, was wer zu zahlen hat
    # wenn derjenige gerade geburtstag hat, ist die bestellung 20% günstiger


def create_new_prices():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Preise'
    ws['A1'].value = 'Gericht'
    ws['B1'].value = 'Preis'
    for i in range(0, len(Dishes)):
        ws["A" + str(i + 2)].value = Dishes[i]
        ws["B" + str(i + 2)].value = random_values.float_value()
    name_prices = "prices.xlsx"
    wb.save(name_prices)
    print("created: " + os.path.abspath(name_prices))


def create_new_orders():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Bestellungen'
    ws['A1'].value = 'Vorname'
    ws['B1'].value = 'Nachname'
    ws['C1'].value = 'Geburtsdatum'
    ws["D1"].value = "Gericht"
    ws["E1"].value = "Anzahl"
    for i in range(2, 60):
        ws['A' + str(i)] = random_values.first_name()
        ws['B' + str(i)].value = random_values.last_name()
        ws['C' + str(i)].value = random_values.date(today=random_values.boolean_value())
        ws['D' + str(i)].value = Dishes[random_values.int_value(0, len(Dishes) - 1)]
        ws['E' + str(i)].value = random_values.int_value()
    name_orders = "orders.xlsx"
    wb.save(name_orders)
    print("created: " + os.path.abspath(name_orders))


gui = tk.Tk()

file_name_input_value = tk.StringVar()

create_new_orders_button = tk.Button(gui, text="Bestellungen neu erstellen", command=create_new_orders)
create_new_orders_button.pack()

create_new_prices_button = tk.Button(gui, text="Preise neu erstellen", command=create_new_prices)
create_new_prices_button.pack()

result_label = tk.Label(gui, text="Dateiname für Ergebnis")
result_label.pack()

file_name_text_box = tk.Entry(gui, textvariable=file_name_input_value)
file_name_text_box.insert(0, "result.xlsx")
file_name_text_box.pack()

create_result_button = tk.Button(gui, text="Berechne Ergebnis", command=create_result)
create_result_button.pack()

gui.mainloop()