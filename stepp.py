from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


def create_empty_excel_file(file_name):
    workbook = Workbook()
    workbook.save(file_name)
    return file_name


def load_excel_file(name):
    return load_workbook(name)


def fill_data(workbook):
    ws = workbook.active
    ws.title = 'Geburtsdatum'
    ws['A1'].value = 'Name'
    ws['A2'].value = 'in Name'
    ws['B1'].value = 'Geburtsdatum'
    ws['B2'].value = 'jeden tag :D'
    ws['C1'].value = 'Aktuelle Zeit'
    now = datetime.now()
    current_time = now.strftime('%H:%M:%S')
    ws['C2'].value = current_time


filenames = ["Bestellungen.xlsx", "Buchungen.xlsx", "Test3.xlsx"]
for file_name in filenames:
    create_empty_excel_file(file_name)
    wb = load_excel_file(file_name)
    fill_data(wb)
    wb.save(file_name)
    print('saved: '+ os.path.abspath(file_name))