from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
#from openpyxl.worksheet.worksheet import Worksheet
#import os

wb = Workbook()
ws = wb
def create_empty_excel_file():  
    a = input('Bitte geben Sie einen Dateinamen ein\n')
    b = a+'.xlsx'
    wb.save(b)
    return(b)
#    print('path: '+os.path.abspath(b))
#     if not os.path.exists(b):    
#        wb.save(b)
#        print('new Workbook saved')
#    else:
#        print('use existing Workbook')

def load_excel_file():
    file_name='Test.xlsx'
    return load_workbook(file_name)

def row_and_col_writer():
    ws = wb.active
    ws.title ='Geburtsdatum'
    ws['A1'].value = 'Name'
    ws['B1'].value = 'Geburtsdatum'  
    ws['A2'].value = 'Iche'
    ws['B2'].value = 'heute'
    ws['C1'].value = 'Irgendwas'
    ws['C2'].value = 'Toll'
    #ws['B100'].value = 'Irgendwas'
    ws['A50'].value = 'Toll'
    wb.save('Test.xlsx')

def row_and_col_reader():
    ws = wb.active
    for row in range(1,3):
        r = ws['A' + str(row)].value
        for col in range(1,3):
            z = ws['A' + str(col)].value
            print(r, z)

def row_and_col_reader_2():
    print("with array")
    ws = wb.active
    for row in range(1,3):
        row_index = str(row)
        for col_index in ['A', 'B', 'C']:
            v = ws[col_index+row_index].value
            print("["+col_index+","+row_index+"] = "+v)

def internet_solution():
    wb.active
    last_row = (wb.active.max_row+1)
    new_row = [cell.value for cell in wb.active[last_row]]
    print(last_row)
    print(new_row)
    if new_row is None:
            new_row['A' + str(last_row)].value = 'Eintrag'
    for i in range(1, wb.active.max_row+1):
        row = [cell.value for cell in wb.active[i]]
        print("Zeile "+str(i), row)
            
wb.save('Test.xlsx')           

create_empty_excel_file()
load_excel_file()
row_and_col_writer()
row_and_col_reader()
row_and_col_reader_2()
internet_solution()
