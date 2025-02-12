import json
import logging
import os

from openpyxl import Workbook

from cards.mapper import json_to_card, json_to_card_set
from downloader import download_data_from_url_to_file
from logging.config import listen

log = logging.getLogger(__name__)


def map_json_file_to_domain_object(local_json_file, mapping_function):
    f = open(local_json_file)
    entire_json = json.load(f)
    objects = set()
    for json_object in entire_json:
        card = mapping_function(json_object)
        objects.add(card)
    f.close()
    return objects


def remove_default_sheet(wb):
    del wb["Sheet"]


def write_cards_to_excel(cards, card_sets, target):
    wb = Workbook()
    current_row_for_card_set = {card_set.set_id: 2 for card_set in card_sets}
    for card in cards:
        sheet_name = card.set_id + " " + str(card.set_num)
        #sheet_name = str(card.set_num) + " " + card.set_id
        #sheet_name = card.set_id
        index_nr = card.set_num
        if sheet_name not in wb.get_sheet_names():
            ws = wb.create_sheet(sheet_name, index_nr)
            add_header_to_sheet(ws)
            wb.get_sheet_names()
        else:
            ws = wb.get_sheet_by_name(sheet_name)
        row = current_row_for_card_set[card.set_id]
        add_values_to_cell(card, row, ws)
        current_row_for_card_set[card.set_id] = row + 1
    remove_default_sheet(wb)
    wb.save(target)
    #sortieren_excl_tabs(wb)
    #sortieren_excel_sheet (ws)
    log.info("wrote cards to %s and %s", os.path.abspath(target), "Hase")


def sortieren_excl_tabs(wb):
    liste = wb.sheetnames
    Anzahl_Elemente = len(liste)
    Index_derliste = liste[-2]
    print ("Anzahl der Elemente" ,Anzahl_Elemente)
    print ("Index_derliste" ,Index_derliste)
    print("Namen: " ,liste)
    wb._sheets.sort(key = lambda ws: ws.title[0])
    print ('sortierte Liste', wb.sheetnames)
    wb.save("cards.xslx")

def add_header_to_sheet(ws):
    ws.cell(row=1, column=1, value="abilities")
    ws.cell(row=1, column=2, value="artist")
    ws.cell(row=1, column=3, value="body_text")
    ws.cell(row=1, column=4, value="card_number")
    ws.cell(row=1, column=5, value="card_variants")
    ws.cell(row=1, column=6, value="classifications")
    ws.cell(row=1, column=7, value="color")
    ws.cell(row=1, column=8, value="cost")
    ws.cell(row=1, column=9, value="flavor_text")
    ws.cell(row=1, column=10, value="image_url")
    ws.cell(row=1, column=11, value="inkable")
    ws.cell(row=1, column=12, value="lore")
    ws.cell(row=1, column=13, value="move_cost")
    ws.cell(row=1, column=14, value="name")
    ws.cell(row=1, column=15, value="rarity")
    ws.cell(row=1, column=16, value="set_id")
    ws.cell(row=1, column=17, value="set_name")
    ws.cell(row=1, column=18, value="set_num")
    ws.cell(row=1, column=19, value="strength")
    ws.cell(row=1, column=20, value="card_type")
    ws.cell(row=1, column=21, value="willpower")


def add_values_to_cell(card, row, ws):
    ws.cell(row, column=1, value=card.abilities)
    ws.cell(row, column=2, value=card.artist)
    ws.cell(row, column=3, value=card.body_text)
    ws.cell(row, column=4, value=card.card_number)
    ws.cell(row, column=5, value=card.card_variants)
    ws.cell(row, column=6, value=card.classifications)
    ws.cell(row, column=7, value=card.color)
    ws.cell(row, column=8, value=card.cost)
    ws.cell(row, column=9, value=card.flavor_text)
    ws.cell(row, column=10, value=card.image_url)
    ws.cell(row, column=11, value=card.inkable)
    ws.cell(row, column=12, value=card.lore)
    ws.cell(row, column=13, value=card.move_cost)
    ws.cell(row, column=14, value=card.name)
    ws.cell(row, column=15, value=card.rarity)
    ws.cell(row, column=16, value=card.set_id)
    ws.cell(row, column=17, value=card.set_name)
    ws.cell(row, column=18, value=card.set_num)
    ws.cell(row, column=19, value=card.strength)
    ws.cell(row, column=20, value=card.card_type)
    ws.cell(row, column=21, value=card.willpower)
    
def sortieren_excel_sheet (ws):
    Spalte_cardnum =ws.cell(row =1, column = 4)
    len(Spalte_cardnum)
    print('Hier wird Len ausgegeben', len)


def handle():
    cards_in_local_file_as_json = download_data_from_url_to_file("https://api.lorcana-api.com/cards/all", "cards/data/cards.json")
    cards = map_json_file_to_domain_object(cards_in_local_file_as_json, json_to_card)
    card_sets_in_local_file_as_json = download_data_from_url_to_file("https://api.lorcana-api.com/sets/all", "cards/data/card_sets.json")
    card_sets = map_json_file_to_domain_object(card_sets_in_local_file_as_json, json_to_card_set)
    write_cards_to_excel(cards, card_sets, "cards/data/cards.xlsx")
